"""Tests for the Excel Faker value fill + openpyxl renderer."""

from __future__ import annotations

import base64
import threading
import time
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from document_gen.generators.excel_gen import (
    ExcelGenerator,
    fill_table_values,
    render_excel_doc,
)
from document_gen.models import (
    BorderStyle,
    Cell,
    CellStyle,
    Column,
    DocSchema,
    ExcelDoc,
    FigurePlacement,
    FigureSpec,
    Sheet,
    Table,
)

# 1x1 transparent PNG (enough for openpyxl's Image to parse dimensions).
_TINY_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJ"
    "AAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
)


def _doc(sheets: list[Sheet], **kwargs) -> ExcelDoc:
    """Build an :class:`ExcelDoc` around the given sheets."""
    return ExcelDoc(
        doc_schema=DocSchema(seed_prompt="test", sheets=[s.name for s in sheets]),
        sheets=sheets,
        created=datetime(2024, 1, 2, 3, 4, 5, tzinfo=timezone.utc),
        **kwargs,
    )


def _table(
    columns: list[Column],
    num_row: int,
    upper_left_position: str = "A1",
    **kwargs,
) -> Table:
    return Table(
        columns=columns,
        num_row=num_row,
        upper_left_position=upper_left_position,
        table_label="test table",
        **kwargs,
    )


def _column(
    headers: list[str] | None = None,
    data_type: str = "str",
    faker_field: str | None = None,
    **kwargs,
) -> Column:
    return Column(
        headers=[Cell(value=h) for h in (headers or ["H"])],
        data_type=data_type,
        faker_field=faker_field,
        **kwargs,
    )


# ---------------------------------------------------------------------------
# fill_table_values
# ---------------------------------------------------------------------------


class TestFillTableValues:
    def test_deterministic_same_seed(self):
        def build():
            doc = _doc(
                [
                    Sheet(
                        name="Data",
                        tables=[
                            _table(
                                [
                                    _column(faker_field="pyint"),
                                    _column(faker_field="name"),
                                ],
                                num_row=5,
                            )
                        ],
                    )
                ]
            )
            return fill_table_values(doc, seed=42)

        a, b = build(), build()
        for col_a, col_b in zip(
            a.sheets[0].tables[0].columns, b.sheets[0].tables[0].columns
        ):
            assert [c.value for c in col_a.cells] == [c.value for c in col_b.cells]

    def test_different_seed_different_values(self):
        def build(seed: int):
            doc = _doc(
                [
                    Sheet(
                        name="Data",
                        tables=[_table([_column(faker_field="pyint")], num_row=20)],
                    )
                ]
            )
            return fill_table_values(doc, seed=seed)

        a = [c.value for c in build(1).sheets[0].tables[0].columns[0].cells]
        b = [c.value for c in build(2).sheets[0].tables[0].columns[0].cells]
        assert a != b

    def test_type_coercion_per_data_type(self):
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[
                        _table(
                            [
                                _column(data_type="int"),  # default pyint
                                _column(data_type="float"),  # default pyfloat
                                _column(
                                    data_type="datetime", faker_field="date_this_year"
                                ),
                                _column(data_type="str", faker_field="words"),
                                _column(
                                    data_type="str", faker_field="credit_card_number"
                                ),
                            ],
                            num_row=3,
                        )
                    ],
                )
            ]
        )
        fill_table_values(doc, seed=7)
        int_col, float_col, dt_col, words_col, cc_col = doc.sheets[0].tables[0].columns
        assert all(type(c.value) is int for c in int_col.cells)
        assert all(isinstance(c.value, float) for c in float_col.cells)
        assert all(isinstance(c.value, datetime) for c in dt_col.cells)
        assert all(isinstance(c.value, str) and " " in c.value for c in words_col.cells)
        assert all(isinstance(c.value, str) for c in cc_col.cells)

    def test_default_faker_field_mapping(self):
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[
                        _table(
                            [
                                _column(data_type="int"),
                                _column(data_type="float"),
                                _column(data_type="datetime"),
                                _column(data_type="str"),
                            ],
                            num_row=2,
                        )
                    ],
                )
            ]
        )
        fill_table_values(doc, seed=11)
        columns = doc.sheets[0].tables[0].columns
        assert all(type(c.value) is int for c in columns[0].cells)
        assert all(isinstance(c.value, float) for c in columns[1].cells)
        assert all(isinstance(c.value, datetime) for c in columns[2].cells)
        assert all(isinstance(c.value, str) for c in columns[3].cells)

    def test_date_only_fields_coerced_to_date(self):
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[
                        _table(
                            [
                                _column(data_type="date", faker_field="date_of_birth"),
                                _column(data_type="date", faker_field="date"),
                            ],
                            num_row=2,
                        )
                    ],
                )
            ]
        )
        fill_table_values(doc, seed=13)
        columns = doc.sheets[0].tables[0].columns
        assert all(isinstance(c.value, date) for c in columns[0].cells)
        assert all(isinstance(c.value, date) for c in columns[1].cells)

    def test_llm_cell_value_overrides_preserved(self):
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[
                        _table(
                            [_column(cells=[Cell(value="keep me"), Cell()])],
                            num_row=3,
                        )
                    ],
                )
            ]
        )
        fill_table_values(doc, seed=19)
        values = [c.value for c in doc.sheets[0].tables[0].columns[0].cells]
        assert values[0] == "keep me"
        assert values[1] is not None
        assert values[1] != "keep me"
        assert values[2] is not None

    def test_cells_beyond_num_row_dropped(self):
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[
                        _table(
                            [
                                _column(
                                    cells=[
                                        Cell(value="a"),
                                        Cell(value="b"),
                                        Cell(value="c"),
                                    ]
                                )
                            ],
                            num_row=2,
                        )
                    ],
                )
            ]
        )
        fill_table_values(doc, seed=23)
        column = doc.sheets[0].tables[0].columns[0]
        assert len(column.cells) == 2
        assert [c.value for c in column.cells] == ["a", "b"]

    def test_returns_same_doc(self):
        doc = _doc([Sheet(name="Data")])
        assert fill_table_values(doc, seed=1) is doc

    def test_concurrent_calls_do_not_share_rng(self):
        """Concurrent fills must each reproduce their single-threaded baseline."""

        def build():
            return _doc(
                [
                    Sheet(
                        name="Data",
                        tables=[_table([_column(faker_field="name")], num_row=10)],
                    )
                ]
            )

        def values(doc: ExcelDoc) -> list:
            return [c.value for c in doc.sheets[0].tables[0].columns[0].cells]

        baseline_a = values(fill_table_values(build(), seed=1))
        baseline_b = values(fill_table_values(build(), seed=2))

        stop = threading.Event()
        mismatches: list[int] = []

        def worker(seed: int, baseline: list) -> None:
            i = 0
            while not stop.is_set():
                if values(fill_table_values(build(), seed=seed)) != baseline:
                    mismatches.append(i)
                i += 1

        threads = [
            threading.Thread(target=worker, args=(1, baseline_a)),
            threading.Thread(target=worker, args=(2, baseline_b)),
        ]
        for t in threads:
            t.start()
        time.sleep(1.0)
        stop.set()
        for t in threads:
            t.join()
        assert mismatches == []


# ---------------------------------------------------------------------------
# render_excel_doc
# ---------------------------------------------------------------------------


class TestRenderExcelDoc:
    def test_multi_row_headers_and_vertical_merge(self, tmp_path: Path):
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[
                        _table(
                            [
                                _column(headers=["Group", "Sub A"]),
                                _column(headers=["Solo"]),
                            ],
                            num_row=2,
                        )
                    ],
                )
            ]
        )
        fill_table_values(doc, seed=1)
        path = render_excel_doc(doc, tmp_path / "out.xlsx")
        ws = openpyxl.load_workbook(path)["Data"]
        assert ws["A1"].value == "Group"
        assert ws["A2"].value == "Sub A"
        assert ws["B1"].value == "Solo"
        # Column B has fewer header cells -> merged across the 2 header rows.
        assert "B1:B2" in [str(r) for r in ws.merged_cells.ranges]
        # Data starts at row 3.
        assert ws["A3"].value is not None

    def test_style_cascade(self, tmp_path: Path):
        table = _table(
            [
                _column(),  # no column style -> table_style
                _column(style=CellStyle(fill_color="#222222")),  # column style
                _column(),  # cell style set below
            ],
            num_row=1,
            table_style=CellStyle(fill_color="#111111"),
            header_style=CellStyle(bold=True),
        )
        table.columns[2].cells = [
            Cell(value="x", style=CellStyle(fill_color="#333333"))
        ]
        doc = _doc([Sheet(name="Data", tables=[table])])
        path = render_excel_doc(doc, tmp_path / "out.xlsx")
        ws = openpyxl.load_workbook(path)["Data"]
        assert ws["A1"].font.bold is True  # header_style
        assert ws["A2"].fill.start_color.rgb.endswith("111111")
        assert ws["B2"].fill.start_color.rgb.endswith("222222")
        assert ws["C2"].fill.start_color.rgb.endswith("333333")

    def test_borders_fills_fonts_number_formats(self, tmp_path: Path):
        style = CellStyle(
            font_color="#FF0000",
            fill_color="00FF00",
            bold=True,
            italic=True,
            font_size=12,
            number_format="#,##0.00",
            alignment="right",
            wrap_text=True,
            border=BorderStyle(style="thick", color="#123456", sides=["top", "left"]),
        )
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[_table([_column(style=style)], num_row=1)],
                )
            ]
        )
        path = render_excel_doc(doc, tmp_path / "out.xlsx")
        ws = openpyxl.load_workbook(path)["Data"]
        cell = ws["A2"]
        assert cell.font.color.rgb.endswith("FF0000")
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.font.size == 12
        assert cell.fill.start_color.rgb.endswith("00FF00")
        assert cell.number_format == "#,##0.00"
        assert cell.alignment.horizontal == "right"
        assert cell.alignment.wrap_text is True
        assert cell.border.top.style == "thick"
        assert cell.border.top.color.rgb.endswith("123456")
        assert cell.border.left.style == "thick"
        assert getattr(cell.border.bottom, "style", None) is None
        assert getattr(cell.border.right, "style", None) is None

    def test_loose_cells_with_merge_range(self, tmp_path: Path):
        doc = _doc(
            [
                Sheet(
                    name="Notes",
                    cells=[
                        Cell(
                            position="B5",
                            value="A paragraph annotation.",
                            merge_range="B5:E8",
                        ),
                        Cell(value="no position -> skipped"),
                    ],
                )
            ]
        )
        path = render_excel_doc(doc, tmp_path / "out.xlsx")
        ws = openpyxl.load_workbook(path)["Notes"]
        assert ws["B5"].value == "A paragraph annotation."
        assert "B5:E8" in [str(r) for r in ws.merged_cells.ranges]
        assert ws["B5"].alignment.wrap_text is True

    def test_hidden_sheets(self, tmp_path: Path):
        doc = _doc(
            [
                Sheet(name="Visible"),
                Sheet(name="Secret", hidden=True),
            ]
        )
        path = render_excel_doc(doc, tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb["Visible"].sheet_state == "visible"
        assert wb["Secret"].sheet_state == "hidden"

    def test_sheet_name_sanitization(self, tmp_path: Path):
        long_name = "x" * 40
        doc = _doc(
            [
                Sheet(name="Bad[]:*?/\\Name"),
                Sheet(name=long_name),
            ]
        )
        path = render_excel_doc(doc, tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert "BadName" in wb.sheetnames
        assert long_name[:31] in wb.sheetnames
        assert len(wb.sheetnames) == 2

    def test_figure_anchoring(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
        calls: list[FigureSpec] = []

        def fake_render(spec: FigureSpec) -> bytes:
            calls.append(spec)
            return _TINY_PNG

        monkeypatch.setattr("document_gen.figures.render_figure_png", fake_render)
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    figures=[FigurePlacement(index=1, anchor="A12")],
                )
            ]
        )
        figures = [FigureSpec(kind="bar", title="Fig", labels=["a"], series=[])]
        path = render_excel_doc(doc, tmp_path / "out.xlsx", figures=figures)
        ws = openpyxl.load_workbook(path)["Data"]
        assert len(ws._images) == 1
        assert calls == figures

    def test_figure_index_out_of_range_skipped(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ):
        monkeypatch.setattr(
            "document_gen.figures.render_figure_png", lambda spec: _TINY_PNG
        )
        doc = _doc(
            [Sheet(name="Data", figures=[FigurePlacement(index=9, anchor="A1")])]
        )
        path = render_excel_doc(doc, tmp_path / "out.xlsx", figures=[])
        ws = openpyxl.load_workbook(path)["Data"]
        assert ws._images == []

    def test_workbook_properties(self, tmp_path: Path):
        doc = _doc(
            [Sheet(name="S")],
            title="My Report",
            creator="Acme Inc",
            keywords=["k1", "k2"],
        )
        path = render_excel_doc(doc, tmp_path / "sub" / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb.properties.title == "My Report"
        assert wb.properties.creator == "Acme Inc"
        assert wb.properties.keywords == "k1 k2"
        assert wb.properties.created.year == 2024

    def test_creates_parent_dirs_and_returns_path(self, tmp_path: Path):
        doc = _doc([Sheet(name="S")])
        target = tmp_path / "a" / "b" / "out.xlsx"
        assert render_excel_doc(doc, target) == target
        assert target.exists()

    def test_empty_doc_renders(self, tmp_path: Path):
        doc = _doc([])
        path = render_excel_doc(doc, tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) == 1


# ---------------------------------------------------------------------------
# ExcelGenerator.generate
# ---------------------------------------------------------------------------


class TestExcelGenerator:
    def test_generate_fills_and_writes(self, tmp_path: Path):
        doc = _doc(
            [
                Sheet(
                    name="Data",
                    tables=[_table([_column(faker_field="name")], num_row=3)],
                )
            ],
            title="Quarterly Report",
        )
        doc.doc_schema.seed = 5
        path = ExcelGenerator().generate(doc, tmp_path)
        assert path == tmp_path / "Quarterly_Report.xlsx"
        ws = openpyxl.load_workbook(path)["Data"]
        assert all(ws.cell(row=r, column=1).value for r in (2, 3, 4))

    def test_generate_unique_paths_for_same_title(self, tmp_path: Path):
        def build() -> ExcelDoc:
            return _doc(
                [Sheet(name="Data", tables=[_table([_column()], num_row=1)])],
                title="Same",
            )

        first = ExcelGenerator().generate(build(), tmp_path)
        second = ExcelGenerator().generate(build(), tmp_path)
        assert first == tmp_path / "Same.xlsx"
        assert second == tmp_path / "Same_1.xlsx"
        assert first.exists() and second.exists()

    def test_generate_rejects_non_excel_doc(self, tmp_path: Path):
        with pytest.raises(TypeError, match="ExcelDoc"):
            ExcelGenerator().generate({"sheets": []}, tmp_path)
