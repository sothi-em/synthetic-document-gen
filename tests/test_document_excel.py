"""Tests for the LLM-driven Excel workbook generation pipeline."""

from __future__ import annotations

import base64
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from document_gen import document_excel, document_query
from document_gen.generators.excel_gen import fill_table_values, render_excel_doc
from document_gen.models import (
    CompanyProfile,
    DocumentType,
    ExcelDoc,
    ExcelPlan,
    FigureExtraction,
    FigureSpec,
    SyntheticCompany,
)

# 1x1 transparent PNG (enough for openpyxl's Image to parse dimensions).
_TINY_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJ"
    "AAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
)


@pytest.fixture(autouse=True)
def isolated(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """Fresh in-memory store; no DOCUMENTS_DIR / settings interference."""
    monkeypatch.setenv("LLM_SETTINGS_PATH", str(tmp_path / "settings.json"))
    monkeypatch.delenv("TINYDB_PATH", raising=False)
    monkeypatch.delenv("DOCUMENTS_DIR", raising=False)
    document_query.reset_db()
    yield tmp_path
    document_query.reset_db()


def _save_company() -> int:
    """Store a company with two report types and return its doc_id."""
    profile = SyntheticCompany(
        name="Acme Corp",
        industry="Finance",
        description="A fictional finance company.",
        headquarters="Boston, Massachusetts",
        size="mid",
    )
    company = CompanyProfile(
        profile=profile,
        reports=[
            DocumentType(
                name="Onboarding Guide",
                category="Guide",
                purpose="New hire orientation",
            ),
            DocumentType(
                name="Quarterly Earnings",
                category="Investor",
                purpose="Quarterly release",
            ),
        ],
        seed=42,
    )
    return document_query.save_company(company)


class FakeBackend:
    """Canned chat backend: plan + ExcelDoc for the structured calls.

    :meth:`query` serves the workbook plan, the spec-only ExcelDoc, or
    the figure-extraction fallback (by model type) and records calls;
    :meth:`complete` serves the markdown draft.
    """

    #: Plan returned for the stage-0 workbook-plan call.
    PLAN = ExcelPlan(
        design_direction="Clean fintech look with navy header bands.",
        palette=["#1F3A5F", "#7A9CC6", "#FFFFFF"],
        sheet_names=["Cover", "Sales"],
        table_density="standard",
        notes="Emphasize the totals row.",
    )

    MARKDOWN = "# Q3 Sales Workbook\n\n## Section\n\n| a | b |\n|---|---|\n| 1 | 2 |\n"

    #: Spec-only ExcelDoc (empty data cells) returned for the styling call.
    EXCEL_DOC: dict[str, Any] = {
        "doc_schema": {"seed_prompt": "Q3 sales workbook", "sheets": ["Sales"]},
        "title": "Q3 Sales Workbook",
        "creator": "Acme Corp",
        "created": "2024-01-15T10:00:00",
        "version": "1.0",
        "keywords": ["sales"],
        "sheets": [
            {
                "name": "Sales",
                "tables": [
                    {
                        "upper_left_position": "A1",
                        "table_label": "Sales table",
                        "num_row": 3,
                        "columns": [
                            {
                                "headers": [{"value": "Name"}],
                                "data_type": "str",
                                "faker_field": "name",
                                "cells": [],
                            },
                            {
                                "headers": [{"value": "Amount"}],
                                "data_type": "float",
                                "faker_field": "pyfloat",
                                "cells": [],
                            },
                        ],
                    }
                ],
            }
        ],
    }

    #: Specs returned by :meth:`query` (the LLM figure-extraction fallback).
    QUERY_FIGURES: list[dict[str, Any]] = [
        {
            "kind": "bar",
            "title": "Trend",
            "labels": ["2021", "2022"],
            "series": [{"name": "Revenue", "values": [100, 120]}],
        }
    ]

    def __init__(self) -> None:
        self.calls: list[dict[str, Any]] = []
        self.query_calls: list[dict[str, Any]] = []

    def complete(
        self,
        prompt: str,
        system: str | None = None,
        deterministic: bool = False,
        seed: int = 0,
        model_name: str | None = None,
        max_tokens: int | None = None,
        thinking: bool = True,
    ) -> str:
        self.calls.append(
            {
                "prompt": prompt,
                "system": system,
                "deterministic": deterministic,
                "seed": seed,
                "model_name": model_name,
                "max_tokens": max_tokens,
                "thinking": thinking,
            }
        )
        return self.MARKDOWN

    def query(
        self,
        prompt: str,
        model: Any,
        deterministic: bool = False,
        seed: int = 0,
        model_name: str | None = None,
        thinking: bool = True,
    ) -> Any:
        self.query_calls.append(
            {
                "prompt": prompt,
                "model": model,
                "deterministic": deterministic,
                "seed": seed,
                "model_name": model_name,
                "thinking": thinking,
            }
        )
        if model is ExcelPlan:
            return self.PLAN
        if model is ExcelDoc:
            return ExcelDoc.model_validate(self.EXCEL_DOC)
        return model(figures=self.QUERY_FIGURES)


def _run(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    backend: FakeBackend,
    **kwargs: Any,
) -> document_excel.ExcelArtifact:
    """Run the pipeline with *backend* canned in and return the artifact."""
    company_id = _save_company()
    monkeypatch.setattr(document_excel, "get_chat_backend", lambda: backend)
    return document_excel.generate_document_excel(
        company_id, "Onboarding Guide", output_dir=tmp_path, **kwargs
    )


def _check_trace(
    artifact: document_excel.ExcelArtifact, backend: FakeBackend, company_id: int
) -> None:
    """Assert the aggregated per-stage trace is complete and DB-stored."""
    trace = artifact.gen_tracing
    assert trace is not None
    stages = trace["stages"]
    # Stage order: plan -> markdown -> figures -> excel -> values -> xlsx.
    assert list(stages) == ["plan", "markdown", "figures", "excel", "values", "xlsx"]
    assert trace["company_id"] == company_id
    assert trace["report"] == "Onboarding Guide"
    assert "started_at" in trace and "finished_at" in trace
    assert trace["total_elapsed_s"] >= 0

    # Plan stage: rendered prompt + structured output (no fallback).
    assert "Acme Corp" in stages["plan"]["prompt"]
    assert stages["plan"]["output"] == FakeBackend.PLAN.model_dump(mode="json")
    assert stages["plan"]["used_default_fallback"] is False

    # Markdown stage: prompt and raw LLM output.
    assert "focus on Q3" in stages["markdown"]["prompt"]
    assert stages["markdown"]["output"] == FakeBackend.MARKDOWN

    # Figures stage: no kinds requested -> nothing extracted.
    assert stages["figures"]["requested_kinds"] == []
    assert stages["figures"]["specs"] == []
    assert stages["figures"]["llm"] is None

    # Excel stage: prompt (embedding the markdown + plan + faker fields)
    # and the full spec-only ExcelDoc JSON.
    assert FakeBackend.MARKDOWN in stages["excel"]["prompt"]
    assert "Clean fintech look with navy header bands." in stages["excel"]["prompt"]
    assert "name" in stages["excel"]["prompt"]  # faker-field whitelist
    # The stored doc is the full spec-only ExcelDoc JSON (round-trips).
    stored = stages["excel"]["excel_doc"]
    expected = ExcelDoc.model_validate(FakeBackend.EXCEL_DOC).model_dump(mode="json")
    assert stored == expected
    # …and it carries no data values.
    assert stored["sheets"][0]["tables"][0]["columns"][0]["cells"] == []

    # Values stage: the Faker seed and the filled value count.
    assert isinstance(stages["values"]["seed"], int)
    assert stages["values"]["filled_values"] == 6  # 2 columns x 3 rows

    # Xlsx stage: the written file.
    assert Path(stages["xlsx"]["path"]) == artifact.xlsx_path
    assert stages["xlsx"]["size_bytes"] > 0


def test_trace_persistence_flag(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    company_id = _save_company()
    monkeypatch.setattr(document_excel, "get_chat_backend", lambda: FakeBackend())

    # Default: the trace is built and returned on the artifact…
    artifact = document_excel.generate_document_excel(
        company_id, "Onboarding Guide", output_dir=tmp_path
    )
    assert artifact.gen_tracing is not None
    # …but the record carries no gen_tracing field.
    record = document_query.list_documents(company_id=company_id)[0]
    assert "gen_tracing" not in record

    # Flagged: the trace is stored on the record.
    artifact = document_excel.generate_document_excel(
        company_id, "Onboarding Guide", output_dir=tmp_path, gen_tracing=True
    )
    record = document_query.list_documents(company_id=company_id)[0]
    assert record["gen_tracing"] == artifact.gen_tracing


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------


class TestGenerateDocumentExcel:
    def test_end_to_end(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        company_id = _save_company()
        backend = FakeBackend()
        monkeypatch.setattr(document_excel, "get_chat_backend", lambda: backend)

        artifact = document_excel.generate_document_excel(
            company_id,
            "Onboarding Guide",
            user_input="focus on Q3",
            output_dir=tmp_path,
            gen_tracing=True,
        )

        assert artifact.company_id == company_id
        assert artifact.report_name == "Onboarding Guide"
        assert artifact.markdown == FakeBackend.MARKDOWN
        assert artifact.xlsx_path.exists()
        assert artifact.xlsx_path.parent == tmp_path
        # File name comes from the markdown's document title.
        assert artifact.xlsx_path.name == "q3_sales_workbook.xlsx"
        # The artifact carries the value-filled doc.
        assert (
            artifact.excel_doc.doc_schema.seed
            == artifact.gen_tracing["stages"]["values"]["seed"]
        )
        assert len(artifact.excel_doc.sheets[0].tables[0].columns[0].cells) == 3

        # One chat call (markdown) and two structured calls (plan, doc).
        assert len(backend.calls) == 1
        assert backend.calls[0]["system"] is None
        assert "focus on Q3" in backend.calls[0]["prompt"]
        assert "Acme Corp" in backend.calls[0]["prompt"]
        assert "Default mode" in backend.calls[0]["prompt"]
        # Glossary is off by default: plain headers, no glossary sheet.
        assert "No glossary sheet" in backend.calls[0]["prompt"]
        assert "No glossary sheet" in backend.query_calls[1]["prompt"]
        assert artifact.gen_tracing["glossary"] is False
        assert len(backend.query_calls) == 2
        assert backend.query_calls[0]["model"] is ExcelPlan
        assert backend.query_calls[1]["model"] is ExcelDoc
        # Non-deterministic (server sampling defaults), seeded; thinking
        # on by default.
        assert all(call["deterministic"] is False for call in backend.calls)
        assert all(call["seed"] == 42 for call in backend.calls)
        assert all(call["thinking"] is True for call in backend.calls)
        assert all(call["thinking"] is True for call in backend.query_calls)
        # The markdown draft is capped.
        assert (
            backend.calls[0]["max_tokens"] == document_excel.MARKDOWN_MAX_TOKENS == 8192
        )

        # The design brief (plan) reached the styling prompt.
        styling_prompt = backend.query_calls[1]["prompt"]
        assert "Planned sheet names: Cover, Sales" in styling_prompt
        assert "Table density: standard" in styling_prompt
        assert "Notes: Emphasize the totals row." in styling_prompt
        assert "None. Do not place any figures." in styling_prompt

        # The workbook renders headers + 3 Faker-filled data rows.
        wb = openpyxl.load_workbook(artifact.xlsx_path)
        ws = wb["Sales"]
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Amount"
        assert all(ws.cell(row=r, column=1).value for r in range(2, 5))
        assert all(
            isinstance(ws.cell(row=r, column=2).value, float) for r in range(2, 5)
        )
        # Workbook properties come from the doc.
        assert wb.properties.title == "Q3 Sales Workbook"
        assert wb.properties.creator == "Acme Corp"

        _check_trace(artifact, backend, company_id)

    def test_trace_rerender_reproduces_workbook(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """The stored trace alone re-renders an identical workbook."""
        artifact = _run(tmp_path, monkeypatch, FakeBackend())
        stages = artifact.gen_tracing["stages"]

        # Rebuild the spec-only doc from the trace, refill with the
        # stored seed, and re-render — no LLM calls, no original doc.
        doc = ExcelDoc.model_validate(stages["excel"]["excel_doc"])
        fill_table_values(doc, stages["values"]["seed"])
        rerendered = tmp_path / "rerendered.xlsx"
        stored_specs = stages["figures"]["specs"]
        render_excel_doc(
            doc,
            rerendered,
            [FigureSpec.model_validate(s) for s in stored_specs] or None,
        )

        def values(path: Path) -> list[list[Any]]:
            ws = openpyxl.load_workbook(path).active
            return [[cell.value for cell in row] for row in ws.iter_rows()]

        assert values(rerendered) == values(artifact.xlsx_path)

    def test_quick_doc_cuts_token_cap_and_thinking(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        backend = FakeBackend()
        _run(tmp_path, monkeypatch, backend, quick_doc=True)
        # The markdown draft cap is cut by 80%…
        assert (
            backend.calls[0]["max_tokens"]
            == int(
                document_excel.MARKDOWN_MAX_TOKENS
                * document_excel.QUICK_DOC_TOKEN_FRACTION
            )
            == 1638
        )
        # …and model thinking is disabled on every LLM call.
        assert all(call["thinking"] is False for call in backend.calls)
        assert all(call["thinking"] is False for call in backend.query_calls)

    def test_simple_sheets_forces_no_figures(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        backend = FakeBackend()
        artifact = _run(
            tmp_path,
            monkeypatch,
            backend,
            simple_sheets=True,
            figure_kinds=["bar", "line"],
        )

        # Figures are forced off: no extraction, no specs, no placements.
        assert artifact.figures == []
        assert artifact.gen_tracing["stages"]["figures"]["requested_kinds"] == []
        assert artifact.gen_tracing["simple_sheets"] is True
        # The plan prompt carries the simple-sheets flag…
        assert "no figures): yes" in backend.query_calls[0]["prompt"]
        assert "Figures to include: none" in backend.query_calls[0]["prompt"]
        # …and the content + styling prompts carry the simple-mode rules.
        assert "Simple sheets mode" in backend.calls[0]["prompt"]
        assert "Simple sheets mode" in backend.query_calls[1]["prompt"]
        # No LLM figure-extraction call at all (plan + doc only).
        assert [call["model"] for call in backend.query_calls] == [
            ExcelPlan,
            ExcelDoc,
        ]

    def test_plan_failure_falls_back_to_default(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        backend = FakeBackend()
        original_query = backend.query

        def failing_query(prompt: str, model: Any, **kwargs: Any) -> Any:
            if model is ExcelPlan:
                raise RuntimeError("LLM down")
            return original_query(prompt, model, **kwargs)

        backend.query = failing_query  # type: ignore[method-assign]
        artifact = _run(tmp_path, monkeypatch, backend)

        # The default plan is used and flagged in the trace…
        assert artifact.gen_tracing["stages"]["plan"]["used_default_fallback"] is True
        default = document_excel._DEFAULT_EXCEL_PLAN
        assert artifact.gen_tracing["stages"]["plan"]["output"] == (
            default.model_dump(mode="json")
        )
        # …and its design brief reached the styling prompt (the failed
        # plan call is not recorded, so the styling call is first).
        assert "Professional, neutral workbook styling." in (
            backend.query_calls[0]["prompt"]
        )
        # The run still completes.
        assert artifact.xlsx_path.exists()

    def test_figure_extraction_heuristic_and_llm_fallback(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.setattr(
            "document_gen.figures.render_figure_png", lambda spec: _TINY_PNG
        )
        backend = FakeBackend()
        # No fenced blocks: every requested kind goes to the LLM fallback.
        backend.MARKDOWN = (
            "# Q3 Sales Workbook\n\n## Section\n\n| a | b |\n|---|---|\n| 1 | 2 |\n"
        )
        # The LLM anchors the figure below the table.
        backend.EXCEL_DOC = {
            **FakeBackend.EXCEL_DOC,
            "sheets": [
                {
                    **FakeBackend.EXCEL_DOC["sheets"][0],
                    "figures": [{"index": 1, "anchor": "A5"}],
                }
            ],
        }
        artifact = _run(
            tmp_path,
            monkeypatch,
            backend,
            figure_kinds=["bar", "line"],
            gen_tracing=True,
        )

        # Plan + doc + one figure-extraction call (the missing kinds).
        assert [call["model"] for call in backend.query_calls] == [
            ExcelPlan,
            FigureExtraction,
            ExcelDoc,
        ]
        assert backend.query_calls[1]["prompt"].count(backend.MARKDOWN) >= 1
        assert "bar, line" in backend.query_calls[1]["prompt"]
        assert [f.kind for f in artifact.figures] == ["bar"]
        # The extracted figure is listed in the styling prompt with its
        # 1-based placement index.
        assert "Figure 1: Trend (bar) — index 1" in backend.query_calls[2]["prompt"]
        # The figures trace captures the LLM fallback.
        llm_trace = artifact.gen_tracing["stages"]["figures"]["llm"]
        assert llm_trace is not None
        assert len(llm_trace["output"]) == 1
        # The figure PNG is anchored in the rendered workbook.
        ws = openpyxl.load_workbook(artifact.xlsx_path).active
        assert len(ws._images) == 1

    def test_error_cases(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setattr(document_excel, "get_chat_backend", lambda: FakeBackend())
        with pytest.raises(ValueError, match="not found"):
            document_excel.generate_document_excel(999, "Onboarding Guide")
        company_id = _save_company()
        with pytest.raises(ValueError, match="output directory"):
            document_excel.generate_document_excel(company_id, "Onboarding Guide")

    def test_output_dir_override_and_collision_suffix(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        company_id = _save_company()
        monkeypatch.setattr(document_excel, "get_chat_backend", lambda: FakeBackend())
        # The explicit output_dir wins over the configured default.
        target = tmp_path / "custom"
        first = document_excel.generate_document_excel(
            company_id, "Onboarding Guide", output_dir=target
        )
        assert first.xlsx_path.parent == target
        # A name collision gets a numeric suffix.
        second = document_excel.generate_document_excel(
            company_id, "Onboarding Guide", output_dir=target
        )
        assert first.xlsx_path != second.xlsx_path
        assert second.xlsx_path.name.endswith("_1.xlsx")

    def test_file_name_falls_back_to_report_name(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        backend = FakeBackend()
        backend.MARKDOWN = "No heading here, just prose.\n"
        artifact = _run(tmp_path, monkeypatch, backend)
        assert artifact.xlsx_path.name == "onboarding_guide.xlsx"

    def test_simple_sheets_post_check_warns(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        backend = FakeBackend()
        # The LLM ignores the simple-mode rules: 5 sheets + a figure.
        backend.EXCEL_DOC = {
            **FakeBackend.EXCEL_DOC,
            "sheets": [
                {
                    "name": f"Sheet {i}",
                    "tables": [],
                    "figures": [{"index": 1, "anchor": "A1"}] if i == 0 else [],
                }
                for i in range(5)
            ],
        }
        with caplog.at_level("WARNING", logger="document_gen.document_excel"):
            _run(tmp_path, monkeypatch, backend, simple_sheets=True)
        warnings = [r.message for r in caplog.records if r.levelname == "WARNING"]
        assert any("5 sheets" in message for message in warnings)
        assert any("figure placements" in message for message in warnings)

    def test_styling_failure_propagates(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        backend = FakeBackend()
        original_query = backend.query

        def failing_query(prompt: str, model: Any, **kwargs: Any) -> Any:
            if model is ExcelDoc:
                raise RuntimeError("LLM down")
            return original_query(prompt, model, **kwargs)

        backend.query = failing_query  # type: ignore[method-assign]
        with pytest.raises(RuntimeError, match="LLM down"):
            _run(tmp_path, monkeypatch, backend)


class TestModeText:
    """The <mode> slot: glossary option on/off, simple vs default mode."""

    def test_default_glossary_off(self) -> None:
        text = document_excel._mode_text(simple_sheets=False, glossary=False)
        assert "Default mode" in text
        assert "No glossary sheet" in text
        assert "Glossary sheet:" not in text

    def test_default_glossary_on(self) -> None:
        text = document_excel._mode_text(simple_sheets=False, glossary=True)
        assert "Default mode" in text
        assert "## Glossary" in text
        assert "TOTAL_SV" in text
        # Abbreviated terms are kept sparse and readable.
        assert "sparingly" in text
        assert "at least 4 characters" in text

    def test_simple_sheets(self) -> None:
        text = document_excel._mode_text(simple_sheets=True, glossary=False)
        assert "Simple sheets mode" in text
        assert "No glossary sheet" in text


class TestGlossaryOption:
    """The glossary checkbox flows through plan + content + styling prompts."""

    def test_glossary_on_prompts_and_trace(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        backend = FakeBackend()
        artifact = _run(tmp_path, monkeypatch, backend, glossary=True)
        # Plan prompt carries the option; content + styling prompts carry
        # the glossary instructions.
        plan_prompt = backend.query_calls[0]["prompt"]
        assert (
            "Glossary sheet (a single lookup sheet defining abbreviated terms):\n  yes"
            in plan_prompt
        )
        assert "Glossary sheet:" in backend.calls[0]["prompt"]
        assert "Glossary sheet:" in backend.query_calls[1]["prompt"]
        assert artifact.gen_tracing["glossary"] is True
