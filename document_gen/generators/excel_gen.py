"""Excel (xlsx) file generator using openpyxl.

The pipeline mirrors the PDF generator's two-step split so each stage is
testable in isolation:

1. :func:`fill_table_values` — deterministic Faker value fill for the
   cells the LLM left empty. The LLM transcribes domain-specific values
   (line items, amounts, ratios, terms) from the markdown draft into the
   cells; only generic personal/contact columns (names, addresses, phone
   numbers, etc.) are left empty with a ``faker_field`` spec. This step
   generates those values from a seeded ``Faker`` instance and coerces
   them to the column's ``data_type``. LLM-authored ``Cell.value``
   overrides are never touched.
2. :func:`render_excel_doc` — pure openpyxl renderer. Assumes values are
   already filled; writes one worksheet per ``Sheet`` with tables, loose
   cells, merged ranges, styling, and anchored matplotlib figures.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.workbook import Workbook

from document_gen.models import (
    Cell,
    CellStyle,
    Column,
    ExcelDoc,
    FigureSpec,
    Sheet,
    Table,
)

from .base import FileGenerator, make_output_path

#: Default Faker field per ``data_type`` when ``faker_field`` is unset.
_DEFAULT_FAKER_FIELD: dict[str, str] = {
    "int": "pyint",
    "float": "pyfloat",
    "datetime": "date_time",
}

#: Characters not allowed in Excel sheet names.
_INVALID_SHEET_CHARS = set("[]:*?/\\")

#: Target rendered width (px) for anchored figures.
_FIGURE_WIDTH_PX = 480

#: Column auto-width bounds.
_MIN_COL_WIDTH = 8.0
_MAX_COL_WIDTH = 50.0


# ---------------------------------------------------------------------------
# Faker value fill
# ---------------------------------------------------------------------------


def _faker_field_for(column: Column) -> str:
    """Resolve the Faker field name for *column*.

    Args:
        column: The column spec.

    Returns:
        The whitelisted Faker field name to call.
    """
    if column.faker_field:
        return column.faker_field
    return _DEFAULT_FAKER_FIELD.get(column.data_type, "word")


def _coerce_value(value: Any, data_type: str) -> Any:
    """Coerce a raw Faker return value to the column's ``data_type``.

    Args:
        value: Raw value returned by Faker (str, int, float, ``date``,
            ``datetime``, or ``list[str]`` for ``words``).
        data_type: Target type name (``int``, ``float``, ``datetime``,
            ``date``, or ``str``/other).

    Returns:
        The coerced value. Unparseable values fall back to a safe default
        for the target type (0 / 0.0) or pass through for ``str``.
    """
    if data_type in ("int", "integer"):
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        try:
            return int(value)
        except (TypeError, ValueError):
            digits = re.sub(r"[^0-9]", "", str(value))
            return int(digits) if digits else 0
    if data_type == "float":
        if isinstance(value, float):
            return value
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0
    if data_type == "datetime":
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            try:
                d = date.fromisoformat(str(value)[:10])
                return datetime(d.year, d.month, d.day)
            except ValueError:
                return value
    if data_type == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return date.fromisoformat(str(value)[:10])
        except ValueError:
            return value
    # str (and any unrecognized data_type): stringify.
    if isinstance(value, (list, tuple)):
        return " ".join(str(v) for v in value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value if isinstance(value, str) else str(value)


def fill_table_values(doc: ExcelDoc, seed: int) -> ExcelDoc:
    """Fill every table's data cells with deterministic Faker values.

    Walks every ``Table`` on every sheet in model order; for each empty
    cell, generates a value from ``faker_field`` (or the ``data_type``
    default mapping) and coerces it to ``data_type``.

    Rules:
    - A ``Cell`` whose ``value`` is already set (LLM-authored override)
      is never touched.
    - Cells beyond ``num_row`` are dropped.
    - Deterministic: same ``doc`` + same ``seed`` -> identical values.
    - Thread-safe: each call uses its own Faker RNG, so concurrent
      calls never share (or re-seed) a common random stream.

    Args:
        doc: The Excel document (mutated in place).
        seed: Random seed for the Faker instance.

    Returns:
        The same *doc* with table values filled.
    """
    from faker import Faker

    # A per-call instance with its own RNG: the class-level
    # ``Faker.seed()`` seeds a module-wide shared Random, so concurrent
    # generation jobs would corrupt each other's value streams (and the
    # stored seed would no longer reproduce the workbook).
    faker = Faker()
    faker.seed_instance(seed)
    for sheet in doc.sheets:
        for table in sheet.tables:
            for column in table.columns:
                field = _faker_field_for(column)
                generate = getattr(faker, field)
                cells = list(column.cells)
                for i in range(table.num_row):
                    existing = cells[i] if i < len(cells) else None
                    if existing is not None and existing.value is not None:
                        continue
                    value = _coerce_value(generate(), column.data_type)
                    if existing is not None:
                        cells[i] = existing.model_copy(update={"value": value})
                    else:
                        cells.append(Cell(value=value))
                column.cells = cells[: table.num_row]
    return doc


# ---------------------------------------------------------------------------
# openpyxl rendering
# ---------------------------------------------------------------------------


def _normalize_hex(value: str) -> str:
    """Normalize a hex color (``#`` optional, 3/4/6/8 digits) for openpyxl.

    Args:
        value: Hex color string.

    Returns:
        Uppercase hex without ``#``; 3/4-digit shorthand is expanded to
        6/8 digits.
    """
    body = value.lstrip("#")
    if len(body) in (3, 4):
        body = "".join(ch * 2 for ch in body)
    return body.upper()


def _apply_style(cell: Any, style: CellStyle | None) -> None:
    """Apply a :class:`CellStyle` to an openpyxl cell (no-op if None).

    Args:
        cell: The openpyxl cell to style.
        style: The style to apply (font, fill, number format, alignment,
            border). ``None`` leaves the cell untouched.
    """
    if style is None:
        return
    if any(
        attr is not None
        for attr in (style.font_color, style.bold, style.italic, style.font_size)
    ):
        cell.font = Font(
            color=_normalize_hex(style.font_color) if style.font_color else None,
            bold=style.bold,
            italic=style.italic,
            size=style.font_size,
        )
    if style.fill_color:
        fill = _normalize_hex(style.fill_color)
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    if style.number_format:
        cell.number_format = style.number_format
    if style.alignment is not None or style.wrap_text is not None:
        cell.alignment = Alignment(
            horizontal=style.alignment,
            wrap_text=style.wrap_text,
        )
    if style.border is not None:
        sides = set(style.border.sides)
        color = _normalize_hex(style.border.color) if style.border.color else None
        cell.border = Border(
            top=Side(style=style.border.style, color=color) if "top" in sides else None,
            bottom=(
                Side(style=style.border.style, color=color)
                if "bottom" in sides
                else None
            ),
            left=(
                Side(style=style.border.style, color=color) if "left" in sides else None
            ),
            right=(
                Side(style=style.border.style, color=color)
                if "right" in sides
                else None
            ),
        )


def _sanitize_sheet_name(name: str) -> str:
    """Sanitize a worksheet name: strip ``[]:*?/\\`` and cap at 31 chars.

    Args:
        name: Raw sheet name.

    Returns:
        A valid worksheet title (``"Sheet"`` if nothing remains).
    """
    cleaned = "".join(ch for ch in name if ch not in _INVALID_SHEET_CHARS).strip()
    return cleaned[:31] or "Sheet"


def _track_length(lengths: dict[int, int], col_idx: int, value: Any) -> None:
    """Record the display length of *value* for column auto-width.

    Args:
        lengths: Mutable map of column index -> max content length.
        col_idx: 1-based column index.
        value: The cell value.
    """
    if value is None:
        return
    length = len(str(value))
    if length > lengths.get(col_idx, 0):
        lengths[col_idx] = length


def _render_table(ws: Any, table: Table, lengths: dict[int, int]) -> None:
    """Render one :class:`Table` into worksheet *ws*.

    Header rows equal ``max(len(column.headers))``; a column with fewer
    header cells is merged vertically across the header rows. Style
    cascade for data cells: cell > column > ``table_style``; header
    cells: header cell style > ``header_style``.

    Args:
        ws: The target worksheet.
        table: The table to render.
        lengths: Mutable map of column index -> max content length.
    """
    top, left = coordinate_to_tuple(table.upper_left_position)
    header_rows = max((len(c.headers) for c in table.columns), default=0)
    for j, column in enumerate(table.columns):
        col_idx = left + j
        for i in range(header_rows):
            cell = ws.cell(row=top + i, column=col_idx)
            if i < len(column.headers):
                cell.value = column.headers[i].value
                _track_length(lengths, col_idx, cell.value)
                _apply_style(cell, column.headers[i].style or table.header_style)
            else:
                _apply_style(cell, table.header_style)
        if header_rows > 1 and len(column.headers) < header_rows:
            ws.merge_cells(
                start_row=top,
                start_column=col_idx,
                end_row=top + header_rows - 1,
                end_column=col_idx,
            )
        for i in range(table.num_row):
            cell = ws.cell(row=top + header_rows + i, column=col_idx)
            data_cell = column.cells[i] if i < len(column.cells) else None
            if data_cell is not None:
                cell.value = data_cell.value
                _track_length(lengths, col_idx, cell.value)
            style = (
                data_cell.style
                if data_cell is not None and data_cell.style is not None
                else None
            )
            _apply_style(cell, style or column.style or table.table_style)


def _render_sheet(
    ws: Any,
    sheet: Sheet,
    figures: list[FigureSpec] | None,
) -> None:
    """Render one :class:`Sheet` into worksheet *ws*.

    Args:
        ws: The target worksheet (already named/hidden as needed).
        sheet: The sheet model.
        figures: Figure specs indexed by ``FigurePlacement.index`` (1-based).
    """
    lengths: dict[int, int] = {}
    for table in sheet.tables:
        _render_table(ws, table, lengths)
    for cell in sheet.cells:
        if not cell.position:
            continue
        target = ws[cell.position]
        target.value = cell.value
        _apply_style(target, cell.style)
        if cell.merge_range:
            ws.merge_cells(cell.merge_range)
            target.alignment = Alignment(
                horizontal=(cell.style.alignment if cell.style is not None else None),
                wrap_text=True,
            )
        _track_length(lengths, coordinate_to_tuple(cell.position)[1], cell.value)
    if figures:
        for placement in sheet.figures:
            idx = placement.index - 1
            if not 0 <= idx < len(figures):
                continue
            from document_gen.figures import render_figure_png

            png = render_figure_png(figures[idx])
            image = OpenpyxlImage(io.BytesIO(png))
            if image.width:
                image.height = int(image.height * _FIGURE_WIDTH_PX / image.width)
                image.width = _FIGURE_WIDTH_PX
            ws.add_image(image, placement.anchor)
    for col_idx, length in lengths.items():
        width = max(_MIN_COL_WIDTH, min(float(length) + 2, _MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def render_excel_doc(
    doc: ExcelDoc,
    path: Path,
    figures: list[FigureSpec] | None = None,
) -> Path:
    """Render a value-filled :class:`ExcelDoc` to an ``.xlsx`` file.

    Pure renderer: assumes table values are already filled (see
    :func:`fill_table_values`). One worksheet per ``Sheet`` (name
    sanitized, ``hidden`` sheets get ``sheet_state = "hidden"``); tables
    at ``upper_left_position``; loose ``Sheet.cells`` at ``position`` with
    optional ``merge_range``; figures matched by
    ``FigurePlacement.index`` and anchored at ``anchor`` (scaled to
    ~480 px wide); column widths estimated from content; workbook
    properties from the doc.

    Args:
        doc: The Excel document (values already filled).
        path: Where to write the ``.xlsx`` file (parents created).
        figures: Optional figure specs in placement-index order.

    Returns:
        *path* (the written file).
    """
    wb = Workbook()
    default_ws = wb.active
    if doc.sheets:
        wb.remove(default_ws)
    else:
        default_ws.title = "Sheet"
    for sheet in doc.sheets:
        ws = wb.create_sheet(title=_sanitize_sheet_name(sheet.name))
        if sheet.hidden:
            ws.sheet_state = "hidden"
        _render_sheet(ws, sheet, figures)
    props = wb.properties
    props.title = doc.title or None
    props.creator = doc.creator or None
    if doc.keywords:
        props.keywords = " ".join(doc.keywords)
    props.created = doc.created
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


class ExcelGenerator(FileGenerator):
    """Render :class:`ExcelDoc` models to ``.xlsx`` files via openpyxl."""

    extension = "xlsx"

    def generate(
        self,
        data: Any,
        output_dir: Path,
        figures: list[FigureSpec] | None = None,
    ) -> Path:
        """Fill Faker values then render *data* to an Excel workbook.

        Pipeline: :func:`fill_table_values` (seeded from
        ``doc.doc_schema.seed``) -> :func:`render_excel_doc`. File names
        come from the doc title; collisions get a ``_1``, ``_2``, … suffix.

        Args:
            data: An :class:`ExcelDoc` to render.
            output_dir: Directory to write the file into.
            figures: Optional figure specs for anchored images.

        Returns:
            Path to the written ``.xlsx`` file.

        Raises:
            TypeError: If *data* is not an :class:`ExcelDoc`.
        """
        if not isinstance(data, ExcelDoc):
            raise TypeError(f"ExcelGenerator expects an ExcelDoc, got {type(data)!r}")
        fill_table_values(data, data.doc_schema.seed)
        stem = data.title.strip().replace(" ", "_") or "document"
        path = make_output_path(output_dir, stem, self.extension)
        counter = 1
        while path.exists():
            path = output_dir / f"{stem}_{counter}.{self.extension}"
            counter += 1
        return render_excel_doc(data, path, figures)
